#!/usr/bin/env python3
"""
excel_to_video.py  –  Step 2 of the Script-Editor + Image-Mapper system
========================================================================
Reads the user-edited Excel control panel → produces:
  1.  revised_script.md      (feed to generate_video.py --article)
  2.  image_mapping.json     (para-index → image-number, ready for PARA_TO_IMAGE_MAP)
  3.  Optionally patches PARA_TO_IMAGE_MAP inside generate_video.py in-place

Slot → image conversion:
    Col B value  N  →  image index  N  (direct, 0-based)
    e.g. 0 → image0,  9 → image9

Usage:
    python3 excel_to_video.py --excel script_editor.xlsx
    python3 excel_to_video.py --excel fraud_video.xlsx --patch-video generate_video.py
"""

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


# ── Excel reader ──────────────────────────────────────────────────────────────

def read_excel(xlsx_path: str) -> list[dict]:
    """
    Read 'Script Editor' sheet.  Data starts at row 4.
    Col A = script text,  Col B = image index (0–9) — direct image number,
    as produced by script_to_excel.py dropdown (values 0,1,2,...,9).
    Returns list of {"para_idx": int, "text": str, "image_num": int}
    """
    wb = load_workbook(xlsx_path, data_only=True)

    if "Script Editor" not in wb.sheetnames:
        raise ValueError(
            f"Sheet 'Script Editor' not found in {xlsx_path}.\n"
            "Was this file created by script_to_excel.py?"
        )

    ws = wb["Script Editor"]
    entries = []
    warnings = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True)):
        text     = row[0] if len(row) > 0 else None
        img_val  = row[1] if len(row) > 1 else None

        if not text or not str(text).strip():
            continue

        para_num = len(entries) + 1   # 1-based, for messages

        if img_val is None or str(img_val).strip() == "":
            warnings.append(f"  [WARN] Para {para_num}: Col B is blank — defaulting to image0")
            image_num = 0
        else:
            try:
                image_num = int(img_val)
                if image_num < 0 or image_num > 9:
                    warnings.append(
                        f"  [WARN] Para {para_num}: image index {image_num} out of range 0–9 — defaulting to image0"
                    )
                    image_num = 0
            except (TypeError, ValueError):
                warnings.append(
                    f"  [WARN] Para {para_num}: invalid value '{img_val}' in Col B — defaulting to image0"
                )
                image_num = 0

        entries.append({
            "para_idx":  len(entries),
            "text":      str(text).strip(),
            "image_num": image_num,
            "slot":      img_val,
        })

    for w in warnings:
        print(w)
    if warnings:
        print()

    return entries


# ── Output writers ─────────────────────────────────────────────────────────────

def write_revised_script(entries: list[dict], output_path: str):
    """
    Write revised_script.md in heading-mode format that generate_video.py reads:
        # Para 1
        <body>

        # Para 2
        <body>
        …
    """
    lines = []
    for e in entries:
        lines.append(f"# Para {e['para_idx'] + 1}")
        lines.append(e["text"])
        lines.append("")
    Path(output_path).write_text("\n".join(lines), encoding="utf-8")
    print(f"  ✅  revised_script.md   →  {output_path}")


def write_image_mapping(entries: list[dict], output_path: str) -> dict:
    """
    Write image_mapping.json with two keys:
      para_to_image_map : {para_idx: image_num}  ← drop-in for PARA_TO_IMAGE_MAP
      detailed          : [{para_idx, slot, image_num, text_preview}, …]
    Returns the para_to_image_map dict.
    """
    para_map = {e["para_idx"]: e["image_num"] for e in entries}
    detailed = [
        {
            "para_idx":     e["para_idx"],
            "slot_chosen":  e["slot"],
            "image_num":    e["image_num"],
            "image_file":   f"image{e['image_num']}_<ddmmyy>.jpeg",
            "text_preview": e["text"][:60] + ("…" if len(e["text"]) > 60 else ""),
        }
        for e in entries
    ]
    payload = {"para_to_image_map": para_map, "detailed": detailed}
    Path(output_path).write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"  ✅  image_mapping.json  →  {output_path}")
    return para_map


# ── In-place patcher for generate_video.py ────────────────────────────────────

def patch_generate_video(video_script: str, para_map: dict):
    """
    Replace the PARA_TO_IMAGE_MAP = { … } block inside generate_video.py
    with the mapping derived from the Excel file.  Creates a .bak backup first.
    """
    path = Path(video_script)
    if not path.exists():
        print(f"  [SKIP] '{video_script}' not found — skipping patch")
        return

    source = path.read_text(encoding="utf-8")

    new_lines = ["PARA_TO_IMAGE_MAP = {"]
    for idx in sorted(para_map.keys()):
        new_lines.append(f"    {idx}: {para_map[idx]},   # Para {idx + 1}")
    new_lines.append("    # paragraphs beyond the above → FALLBACK_IMAGE_NUM")
    new_lines.append("}")
    new_block = "\n".join(new_lines)

    pattern = r"PARA_TO_IMAGE_MAP\s*=\s*\{[^}]*\}"
    if not re.search(pattern, source, re.DOTALL):
        print(f"  [WARN] PARA_TO_IMAGE_MAP block not found in '{video_script}' — skipping patch")
        return

    patched = re.sub(pattern, new_block, source, flags=re.DOTALL)
    backup  = path.with_suffix(".py.bak")
    backup.write_text(source, encoding="utf-8")
    path.write_text(patched, encoding="utf-8")
    print(f"  ✅  Patched '{video_script}'  (backup → {backup.name})")


# ── Summary table ──────────────────────────────────────────────────────────────

def print_summary(entries: list[dict]):
    print()
    print(f"  {'Para':>5}  {'Slot':>5}  {'→ Image':>8}  {'Text preview':<52}")
    print("  " + "─" * 74)
    for e in entries:
        preview = e["text"][:50] + ("…" if len(e["text"]) > 50 else "")
        print(f"  {e['para_idx']+1:>5}  {str(e['slot'] or '—'):>5}  image{e['image_num']:<5}  {preview}")
    print()


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description="Excel control panel → revised_script.md + image_mapping.json"
    )
    p.add_argument("--excel",       default="script_editor.xlsx",
                   help="Path to user-edited Excel file")
    p.add_argument("--out-dir",     default=".",
                   help="Directory for output files (default: current dir)")
    p.add_argument("--patch-video", default=None,
                   help="Optional: path to generate_video.py to patch PARA_TO_IMAGE_MAP in-place")
    args = p.parse_args()

    print("=" * 60)
    print("  Excel → Video Assets  (Step 2 of 2)")
    print("=" * 60)
    print(f"  Reading  : {args.excel}")

    entries = read_excel(args.excel)
    print(f"  Rows     : {len(entries)} paragraphs")

    print_summary(entries)

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    write_revised_script(entries, str(out_dir / "revised_script.md"))
    para_map = write_image_mapping(entries, str(out_dir / "image_mapping.json"))

    if args.patch_video:
        patch_generate_video(args.patch_video, para_map)

    print()
    print("  Next step:")
    print(f"    python3 generate_video.py --article {out_dir / 'revised_script.md'}")
    print("=" * 60)


if __name__ == "__main__":
    main()
