#!/usr/bin/env python3
"""
script_to_excel.py  –  Step 1 of the Script-Editor + Image-Mapper system
=========================================================================
Reads any script.md → splits into paragraphs → creates an Excel control panel
with editable script text (Col A) and image-selection dropdowns (Col B).

Dropdown values : 0–9  (direct image index — 0=image0.jpeg … 9=image9.jpeg)
Col B starts BLANK — user must explicitly choose an image for every paragraph.

Usage:
    python3 script_to_excel.py --script script.md --out script_editor.xlsx
    python3 script_to_excel.py --script my_new_video.md --out new_video.xlsx

After editing Excel, run excel_to_video.py to produce revised_script.md
and image_mapping.json.
"""

import argparse
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ── Brand colours ──────────────────────────────────────────────────────────────
CLR_HEADER_BG  = "1F3864"   # deep navy
CLR_HEADER_FG  = "FFFFFF"
CLR_PARA_BG    = "EBF3FB"   # light blue stripe
CLR_PARA_ALT   = "FFFFFF"
CLR_IMG_BLANK  = "FCE4D6"   # orange-red = needs attention (blank cell)
CLR_IMG_FILLED = "FFF2CC"   # amber = filled cell
CLR_IMG_HEADER = "BF8F00"
CLR_ACCENT     = "2E75B6"
CLR_BORDER     = "B8CCE4"
CLR_HELP_BG    = "E2EFDA"
CLR_HELP_FG    = "375623"

NUM_IMAGES     = 10                                              # fallback; overridden dynamically in build_excel


def thin_border(color=CLR_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


# ── Paragraph parser (matches generate_video.py logic exactly) ────────────────

def parse_paragraphs(filepath: str) -> list[dict]:
    """
    Heading-mode  : lines starting with # begin a new section.
    Blank-line mode: blank-line separated blocks, titled Para 1, Para 2 …
    Returns list of {"title": str, "body": str}.
    """
    with open(filepath, encoding="utf-8") as f:
        raw = f.read()

    has_headings = any(ln.lstrip().startswith("#") for ln in raw.splitlines())
    sections = []

    if has_headings:
        lines = raw.split("\n")
        cur_title, cur_body = None, []
        for line in lines:
            line = line.strip()
            if not line:
                continue
            if line.startswith("#"):
                if cur_title:
                    body = " ".join(cur_body).strip()
                    if body:
                        sections.append({"title": cur_title, "body": body})
                cur_title = re.sub(r"^#+\s*", "", line).strip()
                cur_body = []
            else:
                if cur_title is None:
                    cur_title = "Introduction"
                cur_body.append(line)
        if cur_title and cur_body:
            body = " ".join(cur_body).strip()
            if body:
                sections.append({"title": cur_title, "body": body})
    else:
        blocks = [b.strip() for b in re.split(r"\n\s*\n", raw) if b.strip()]
        for i, block in enumerate(blocks, 1):
            lines_txt = [ln.strip() for ln in block.splitlines() if ln.strip()]
            sections.append({"title": f"Para {i}", "body": " ".join(lines_txt)})

    return sections


# ── Excel builder ──────────────────────────────────────────────────────────────

def build_excel(sections: list[dict], output_path: str):
    n        = len(sections)
    # image0…image(n-2) are unique; last para reuses image0 → n-1 distinct images
    n_imgs   = max(n - 1, 1)
    max_img  = n_imgs - 1
    dropdown_str = ",".join(str(i) for i in range(n_imgs))

    wb = Workbook()

    # ── Sheet 1 : Script Editor ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Script Editor"

    # Row 1 – title banner
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value     = "Script Editor & Image Mapper"
    c.font      = Font(name="Arial", bold=True, size=14, color=CLR_HEADER_FG)
    c.fill      = PatternFill("solid", fgColor=CLR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Row 2 – instruction strip
    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value = (
        f"Col A: Edit script text freely.   "
        f"Col B: Image number auto-assigned (0=image0 … {max_img}=image{max_img}).   "
        f"First & last para share image0.   "
        f"Override any row via the ▼ dropdown before running excel_to_video.py."
    )
    c.font      = Font(name="Arial", italic=True, size=9, color=CLR_HELP_FG)
    c.fill      = PatternFill("solid", fgColor=CLR_HELP_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    # Row 3 – column headers
    headers   = ["Script Text (Editable)", f"Image No.  ▼ (0–{max_img})", "Para #", "Section Title / Notes"]
    hdr_fills = [CLR_ACCENT, CLR_IMG_HEADER, CLR_ACCENT, CLR_ACCENT]
    for col_i, (hdr, bg) in enumerate(zip(headers, hdr_fills), 1):
        cell = ws.cell(row=3, column=col_i, value=hdr)
        cell.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
    ws.row_dimensions[3].height = 24

    # Dropdown validation : allow_blank=True so cells start empty without error
    dv = DataValidation(
        type="list",
        formula1=f'"{dropdown_str}"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid selection",
        error=f"Please choose a number between 0 and {max_img}.",
    )
    ws.add_data_validation(dv)

    # Data rows
    for para_idx, sec in enumerate(sections):
        row = para_idx + 4
        bg  = CLR_PARA_BG if para_idx % 2 == 0 else CLR_PARA_ALT

        # Col A – editable script text
        a = ws.cell(row=row, column=1, value=sec["body"])
        a.font      = Font(name="Arial", size=11)
        a.fill      = PatternFill("solid", fgColor=bg)
        a.alignment = Alignment(wrap_text=True, vertical="top")
        a.border    = thin_border()

        # Col B – para_idx maps 1:1 to image number; last para wraps back to image0
        auto_img = para_idx if para_idx < n - 1 else 0
        b = ws.cell(row=row, column=2, value=auto_img)
        b.font      = Font(name="Arial", bold=True, size=12, color="7F3F00")
        b.fill      = PatternFill("solid", fgColor=CLR_IMG_FILLED)
        b.alignment = Alignment(horizontal="center", vertical="center")
        b.border    = thin_border("BF8F00")
        dv.add(b)

        # Col C – para number (informational)
        c = ws.cell(row=row, column=3, value=f"Para {para_idx + 1}")
        c.font      = Font(name="Arial", size=10, color="595959")
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()

        # Col D – section title as hint
        title_hint = sec["title"] if sec["title"] not in (f"Para {para_idx + 1}", "") else ""
        d = ws.cell(row=row, column=4, value=title_hint)
        d.font      = Font(name="Arial", size=10, italic=True, color="595959")
        d.fill      = PatternFill("solid", fgColor=bg)
        d.alignment = Alignment(wrap_text=True, vertical="top")
        d.border    = thin_border()

        approx_lines = max(2, len(sec["body"]) // 90 + 1)
        ws.row_dimensions[row].height = max(40, approx_lines * 18)

    ws.column_dimensions["A"].width = 72
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 32
    ws.freeze_panes = "A4"

    # ── Sheet 2 : Image Reference Card ────────────────────────────────────────
    ws2 = wb.create_sheet("Image Reference")

    ws2.merge_cells("A1:C1")
    c = ws2["A1"]
    c.value     = "Image Reference  (Col B number → image file)"
    c.font      = Font(name="Arial", bold=True, size=13, color=CLR_HEADER_FG)
    c.fill      = PatternFill("solid", fgColor=CLR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    for col_i, hdr in enumerate(
        ["Image No. (Col B)", "Filename"], 1
    ):
        cell      = ws2.cell(row=2, column=col_i, value=hdr)
        cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=CLR_ACCENT)
        cell.alignment = Alignment(horizontal="center")

    for img_num in range(n_imgs):
        r = img_num + 3
        ws2.cell(row=r, column=1, value=img_num)
        ws2.cell(row=r, column=2, value=f"image{img_num}.jpeg")
        fill_color = "EBF3FB" if img_num % 2 == 0 else "FFFFFF"
        for col_i in range(1, 3):
            ws2.cell(row=r, column=col_i).fill = PatternFill("solid", fgColor=fill_color)
        for col_i in range(1, 3):
            ws2.cell(row=r, column=col_i).alignment = Alignment(horizontal="center")

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 28

    # ── Sheet 3 : Instructions ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Instructions")

    steps = [
        ("🎯  How to use this workbook", True, 13, CLR_HEADER_BG, CLR_HEADER_FG),
        ("", False, 10, None, None),
        ("STEP 1 – Edit the script  (Column A)", True, 11, CLR_ACCENT, "FFFFFF"),
        ("  • Click any cell in Column A and edit the Hindi/Hinglish text freely.", False, 10, None, None),
        ("  • Add, remove, or reword sentences.  Text wraps automatically.", False, 10, None, None),
        ("", False, 10, None, None),
        ("STEP 2 – Assign an image number  (Column B)", True, 11, CLR_ACCENT, "FFFFFF"),
        ("  • Every row starts BLANK (orange tint).  You must fill ALL rows.", False, 10, None, None),
        ("  • Click the ▼ dropdown and choose a number from 0 to 9.", False, 10, None, None),
        ("  • 0 → image0.jpeg", False, 10, None, None),
        ("  • 1 → image1.jpeg", False, 10, None, None),
        ("  • …", False, 10, None, None),
        ("  • 9 → image9.jpeg", False, 10, None, None),
        ("  • Multiple paragraphs can use the same image number.", False, 10, None, None),
        ("  • See the 'Image Reference' sheet for the complete table.", False, 10, None, None),
        ("", False, 10, None, None),
        ("STEP 3 – Save the file", True, 11, CLR_ACCENT, "FFFFFF"),
        ("  • File → Save  (keep .xlsx format, any filename you like).", False, 10, None, None),
        ("", False, 10, None, None),
        ("STEP 4 – Generate video assets", True, 11, CLR_ACCENT, "FFFFFF"),
        ("  • Run:  python3 excel_to_video.py --excel <your_file>.xlsx", False, 10, None, None),
        ("  • Output:  revised_script.md  +  image_mapping.json", False, 10, None, None),
        ("  • Then:   python3 generate_video.py --article revised_script.md", False, 10, None, None),
    ]

    for row_i, (text, bold, size, bg, fg) in enumerate(steps, 1):
        ws3.merge_cells(f"A{row_i}:C{row_i}")
        c = ws3[f"A{row_i}"]
        c.value     = text
        c.font      = Font(name="Arial", bold=bold, size=size, color=fg or "000000")
        if bg:
            c.fill  = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws3.row_dimensions[row_i].height = 20

    ws3.column_dimensions["A"].width = 90

    wb.save(output_path)


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description="Convert any script.md into an Excel image-mapping control panel"
    )
    p.add_argument("--script", default="script.md",          help="Input markdown file")
    p.add_argument("--out",    default="script_editor.xlsx",  help="Output Excel filename")
    args = p.parse_args()

    print("=" * 60)
    print("  Script → Excel  (Step 1 of 2)")
    print("=" * 60)

    script_path = Path(args.script)
    if not script_path.exists():
        demo = """# परिचय
आज हम आपको एक महत्वपूर्ण साइबर सुरक्षा विषय के बारे में बताएंगे। यह जानकारी हर भारतीय नागरिक के लिए जरूरी है।

# OTP क्या होता है?
OTP यानी One Time Password एक सुरक्षा कोड होता है। यह आपके बैंक या मोबाइल पर आता है। इसे कभी किसी के साथ शेयर न करें।

# फ्रॉड कैसे होता है?
जालसाज खुद को बैंक अधिकारी बताकर फोन करते हैं। वे आपसे OTP मांगते हैं। एक बार OTP देने के बाद आपका खाता खाली हो सकता है।

# सावधानी बरतें
कोई भी बैंक आपसे फोन पर OTP नहीं मांगता। अनजान कॉल पर कोई जानकारी न दें। तुरंत अपने नजदीकी बैंक शाखा से संपर्क करें।

# क्या करें?
अगर आपके साथ fraud हुआ हो तो 1930 पर कॉल करें। यह साइबर क्राइम हेल्पलाइन है। जितनी जल्दी रिपोर्ट करेंगे, उतना बेहतर।
"""
        script_path.write_text(demo, encoding="utf-8")
        print(f"  [INFO] '{args.script}' not found — demo script created.")

    print(f"  Reading : {args.script}")
    sections = parse_paragraphs(str(script_path))
    print(f"  Found   : {len(sections)} paragraphs")

    build_excel(sections, args.out)

    n = len(sections)
    print(f"  ✅  Excel created : {args.out}")
    print(f"      Paragraphs    : {n}")
    print(f"      Image numbers : 0–{n-2}  (image0.jpeg … image{n-2}.jpeg, last para → image0)")
    print(f"      Col B         : auto-assigned para_idx → image_idx, last para = image0")
    print()
    print("  Next steps:")
    print(f"    1. Open {args.out}")
    print(f"    2. Edit Col A text, override image number (0–{n-2}) in Col B if needed")
    print("    3. Save the file")
    print(f"    4. Run:  python3 excel_to_video.py --excel {args.out}")
    print("=" * 60)


if __name__ == "__main__":
    main()
